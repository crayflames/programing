#! /usr/bin/python3
# -*- coding: utf8 -*-
from openpyxl import Workbook
import datetime
wb=Workbook()
ws=wb.active
ws['A1']=42
ws.append([3,2,3])
ws['A2'] = datetime.datetime.now()
ws2=wb.create_sheet("sheet2",0)
ws2.title='changetitle'
ws2=wb.active
ws2['A1']=42
ws2.append(['aa','bb','cc'])
ws2.append=datetime.datetime.now()
wb.save("/home/roger/Desktop/example.xlsx")