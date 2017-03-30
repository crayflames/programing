#! /usr/bin/python3
# -*- coding: utf8 -*-
from openpyxl import Workbook
import datetime
import requests
from bs4 import BeautifulSoup
res=requests.get('http://benjr.tw/')
wes=BeautifulSoup(res.text,"html.parser")
ss=wes.select(".entry-title")
xx=wes.select(".entry-meta")
wb=Workbook()
ws=wb.active
ws.title='blogtitle'
for i in range(1,len(ss)):
	content=ss[i-1].text.strip()
	time=xx[i-1].text.strip()
	ws.cell(row=i, column=1, value=content)
	ws.cell(row=i, column=2, value=time)
wb.save("/home/roger/Desktop/example2.xlsx")
