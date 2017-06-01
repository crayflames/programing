#! /usr/bin/python3
# -*- coding: utf8 -*-
from openpyxl import load_workbook
wb = load_workbook(filename=r'/home/roger/Desktop/example.xlsx')
sheets = wb.get_sheet_names()
sheets[0]=wb.active
print(sheets[0])
ws=sheets[0]
ws['B13']=12345678
ws.append(['aa','bb','cc'])
ws.append(['ff','ee','dd'])
ws.cell(row=4, column=2, value=10)
for i in range(1,100):
	for j in range(1,10):
		r=int(i)+int(j)
		ws.cell(row=i,column=j,value=r)
print(ws['C10'].value)
wb.save("/home/roger/Desktop/example.xlsx")


